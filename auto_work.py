import time
import openpyxl
import sys
import os
from PyQt5.Qt import QWidget,QApplication
from main import Ui_Form
from work import AutoWork
from PyQt5.QtWidgets import QFileDialog


class Window(QWidget,Ui_Form):
    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.setupUi(self)
        self.file = None

    def clear(self):
        self.textEdit.clear()

    def clear_line_edit(self):
        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()

    def submit(self):
        content_1 = self.lineEdit.text()
        content_2 = self.lineEdit_2.text()
        content_3 = self.lineEdit_3.text()
        content_4 = self.lineEdit_4.text()
        content_5 = self.lineEdit_5.text()
        print(content_1,content_2,content_3,content_4,content_5)
        print(type(content_2))
        if not content_2:
            content_2 = '@'
        if not content_3:
            content_3 = '@'
        if not content_4:
            content_4 = '@'
        if not content_5:
            content_5 = '@'
        self.textEdit.append(content_1+','+content_2+','+content_3+','+content_4+','+content_5)
        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()

    def doit(self):
        if self.file:
            autowork = AutoWork(self.file)
            autowork.main_work()
        else:
            file = 'cmd.xlsx'
            current_directory = os.path.dirname(__file__)
            file_path = os.path.join(current_directory,file)

            if os.path.exists(file_path):
                os.remove(file_path)

            wb = openpyxl.Workbook()
            sheet = wb.active
            content_list = self.textEdit.toPlainText().split('\n')
            print(content_list)
            for i,line in enumerate(content_list):
                i += 2
                line_list = line.split(',')
                for j,value in enumerate(line_list):
                    j += 1
                    if j != 2 and value != '@':
                        print(float(value))
                        if float(value) < 1:
                            sheet.cell(row=i, column=j).value = float(value)
                            print('是小数')
                        else:
                            sheet.cell(row=i,column=j).value = int(value)
                            print('是整数')

                    elif value == '@':
                        continue
                    else:
                        sheet.cell(row=i, column=j).value = value

            wb.save('cmd.xlsx')
            autowork = AutoWork('cmd.xlsx')
            autowork.main_work()

    def export(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        content_list = self.textEdit.toPlainText().split('\n')
        print(content_list)
        for i, line in enumerate(content_list):
            i += 2
            line_list = line.split(',')
            for j, value in enumerate(line_list):
                j += 1
                if j != 2 and value != '@':
                    print(float(value))
                    if float(value) < 1:
                        sheet.cell(row=i, column=j).value = float(value)
                        print('是小数')
                    else:
                        sheet.cell(row=i, column=j).value = int(value)
                        print('是整数')

                elif value == '@':
                    continue
                else:
                    sheet.cell(row=i, column=j).value = value

        min = str(time.localtime().tm_min)
        sec = str(time.localtime().tm_sec)

        wb.save(min + '_' +  sec + 'cmd.xlsx')

    def import_data(self):
        options = QFileDialog.Options()
        file_dialog = QFileDialog(self)
        file_name, _ = file_dialog.getOpenFileName(self, "打开文件", "", "All Files (*);;Text Files (*.txt)",options=options)
        print(file_name)

        content = ''
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        print(sheet.max_row)
        for value in range(2,sheet.max_row+2):
            for col in range(1,8):
                if not sheet.cell(row=value,column=col).value:
                    content += '@' + ','
                else:
                    content += str(sheet.cell(row=value, column=col).value) + ','
            self.textEdit.append(content+'\n')
            content = ''

        self.file = file_name



if __name__ == '__main__':
    app = QApplication(sys.argv)

    w = Window()

    w.show()

    app.exec_()