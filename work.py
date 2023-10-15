import pyautogui as pag
import time
import openpyxl
import pyperclip
import os



class AutoWork():
    def __init__(self,filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.sheet = self.wb.active
        self.data_check()

    def main_work(self):
        i = 2
        while i <= self.sheet.max_row:
            cmdvalue = self.sheet.cell(row=i,column=1).value
            if cmdvalue == 1:
                img = self.sheet.cell(row=i,column=2).value
                print(img)
                if type(self.sheet.cell(row=i,column=4).value) != type(None):
                    retry = self.sheet.cell(row=i,column=4).value
                else:
                    retry = 1
                self.click_picture(img,retry)
            elif cmdvalue == 2:
                localte_x = self.sheet.cell(row=i,column=3).value
                localte_y = self.sheet.cell(row=i,column=4).value
                if type(self.sheet.cell(row=i,column=5).value) != type(None):
                    retry = self.sheet.cell(row=i,column=5).value
                else:
                    retry = 1
                self.click_x_y(localte_x,localte_y,retry)
            elif cmdvalue == 3:
                localte_x = self.sheet.cell(row=i, column=3).value
                localte_y = self.sheet.cell(row=i, column=4).value
                if type(self.sheet.cell(row=i, column=5).value) != type(None):
                    retry = self.sheet.cell(row=i, column=5).value
                else:
                    retry = 1
                self.click_right(localte_x,localte_y,retry)
            elif cmdvalue == 4:
                content = self.sheet.cell(row=i,column=2).value
                if type(self.sheet.cell(row=i,column=5).value) != type(None):
                    retry = self.sheet.cell(row=i,column=5).value
                else:
                    retry = 1
                self.write_content(content,retry)
            elif cmdvalue == 5:
                print('进入等待函数')
                w_time = self.sheet.cell(row=i,column=3).value
                self.waite_time(w_time)
            elif cmdvalue == 6:
                scroll = self.sheet.cell(row=i,column=3).value
                if type(self.sheet.cell(row=i,column=5).value) != type(None):
                    retry = self.sheet.cell(row=i,column=5).value
                else:
                    retry = 1
                self.mouse_scroll(scroll,retry)
            elif cmdvalue == 7:
                if type(self.sheet.cell(row=i,column=5).value) != type(None):
                    retry = self.sheet.cell(row=i,column=5).value
                else:
                    retry = 1
                self.enter(retry)
            elif cmdvalue == 8:
                localte_x = self.sheet.cell(row=i, column=3).value
                localte_y = self.sheet.cell(row=i, column=4).value
                if type(self.sheet.cell(row=i, column=5).value) != type(None):
                    retry = self.sheet.cell(row=i, column=5).value
                else:
                    retry = 1
                self.move(localte_x,localte_y,retry)
            i += 1

    def data_check(self):
        """检查数据是否合法"""
        check_result = True

        if self.sheet.max_row < 2:
            print("配置文件中没有数据，请检查配置文件")
            check_result = False
            return check_result

        i = 2
        while i <= self.sheet.max_row:
            # 检查指令类型是否为数字
            cmdvalue = self.sheet.cell(row=i,column=1).value
            if not isinstance(cmdvalue,int) or (cmdvalue != 1 and cmdvalue != 2 and cmdvalue != 3 and cmdvalue != 4 and cmdvalue != 5 and cmdvalue != 6 and cmdvalue != 7 and cmdvalue != 8):
                print('指令类型不是数字，或者输入的执行范围不在1-8')
                check_result = False
                return check_result
            # 单击图片
            if cmdvalue == 1:
                if not isinstance(self.sheet.cell(row=i,column=2).value,str) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=4).value,int))):
                    print("第%s行第2列或第4列数据有问题" % i)
                    check_result = False
            # 单击坐标
            elif cmdvalue == 2:
                if not isinstance(self.sheet.cell(row=i,column=3).value,int) or not isinstance(self.sheet.cell(row=i,column=4).value,int) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列或第3或第4列数据有问题" % i)
                    check_result = False
            # 右键
            elif cmdvalue == 3:
                if not isinstance(self.sheet.cell(row=i,column=3).value,int) or not isinstance(self.sheet.cell(row=i,column=4).value,int) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列或第3列或第4有多余的数据" % i)
                    check_result = False
            # 输入内容
            elif cmdvalue == 4:
                if not isinstance(self.sheet.cell(row=i,column=2).value,str) or type(self.sheet.cell(row=i,column=5).value) != type(None) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False
            # 时间等待
            elif cmdvalue == 5:
                if not isinstance(self.sheet.cell(row=i,column=3).value,int) or type(self.sheet.cell(row=i,column=4).value) != type(None) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False
            # 滚轮
            elif cmdvalue == 6:
                if not isinstance(self.sheet.cell(row=i,column=2).value,int) or type(self.sheet.cell(row=i,column=3).value) != type(None) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False
            # 回车
            elif cmdvalue == 7:
                if type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int)):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False
            # 移动鼠标
            elif cmdvalue == 8:
                if not isinstance(self.sheet.cell(row=i,column=3).value,int) or not isinstance(self.sheet.cell(row=i,column=4).value,int) or (type(self.sheet.cell(row=i,column=5).value) != type(None) and (not isinstance(self.sheet.cell(row=i,column=5).value,int))):
                    print("第%s行第2列,3列，4列数据有问题" % i)

            i += 1
            return check_result

    def click_picture(self,img,retry):
        if retry == 1:
            while True:
                location = pag.locateCenterOnScreen(os.path.join(os.getcwd(),'picture/'+ img),confidence=0.9)
                if location is not None:
                    pag.click(location.x,location.y,clicks=1,interval=0.2,duration=0.2,button="left")
                    break
                print("没有找到图片，正在重新查找")
                time.sleep(0.5)
        elif retry == -1:
            while True:
                location = pag.locateCenterOnScreen(os.path.join(os.getcwd(),'picture/'+ img), confidence=0.9)
                if location is not None:
                    pag.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button="left")
                print("没有找到图片，正在重新查找")
                time.sleep(0.5)
        elif retry >= 2:
            while retry:
                location = pag.locateCenterOnScreen(os.path.join(os.getcwd(),'picture/'+ img), confidence=0.9)
                if location is not None:
                    pag.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button="left")
                    retry -= 1
                print("没有找到图片，正在重新查找")
                time.sleep(0.5)

    def click_x_y(self,x,y,retry):
        if retry == 1:
            pag.click(x, y, button='left')
            time.sleep(0.5)
        elif retry == -1:
            while True:
                pag.click(x, y, button='left')
                time.sleep(1)
                print('重复')
        elif retry >= 2:
            while True:
                retry -= 1
                pag.click(x, y, button='left')
                time.sleep(1)
                print('重复')

    def click_right(self,x,y,retry):
        if retry == 1:
            pag.click(x, y, button='right')
            time.sleep(0.5)
        elif retry == -1:
            while True:
                pag.click(x, y, button='right')
                time.sleep(1)
                print('重复')
        elif retry >= 2:
            while True:
                retry -= 1
                pag.click(x, y, button='right')
                time.sleep(1)
                print('重复')

    def move(self,x,y,retry):
        if retry == 1:
            pag.moveTo(x,y)
        elif retry == -1:
            while True:
                pag.moveTo(x,y)
        elif retry >= 2:
            while retry:
                pag.moveTo(x,y)
                retry -= 1

    def write_content(self,content,retry):
        pyperclip.copy(content)
        if retry == 1:
            pag.hotkey('ctrl','v')
        elif retry == -1:
            while True:
                pag.hotkey('ctrl', 'v')
        elif retry >= 2:
            while retry:
                pag.hotkey('ctrl', 'v')
                retry -= 1

    def waite_time(self,w_time):
        print('等待%s秒' % w_time)
        time.sleep(w_time)

    def mouse_scroll(self,scroll,retry):
        if retry == 1:
            pag.scroll(scroll)
            print('滑动滚轮')
            time.sleep(0.5)
        elif retry == -1:
            while True:
                pag.scroll(scroll)
                print('滑动滚轮')
                time.sleep(0.5)
        elif retry >= 2:
            while retry:
                pag.scroll(scroll)
                print('滑动滚轮')
                time.sleep(0.5)
                retry -= 1

    def enter(self,retry):
        if retry == 1:
            pag.press('enter')
            time.sleep(0.5)
        elif retry == -1:
            while True:
                pag.press('enter')
                time.sleep(0.5)
        elif retry >= 2:
            while retry:
                pag.press('enter')
                time.sleep(0.5)
                retry -= 1


if __name__ == '__main__':
    a = AutoWork('cmd.xlsx')
    a.main_work()